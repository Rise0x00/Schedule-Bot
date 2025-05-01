from openpyxl import load_workbook
import openpyxl as pd
from openpyxl.utils import get_column_letter

# Строки начала и конца дня недели
monday = (4, 19)
tuesday = (20, 33)
wednesday = (34, 47)
thursday = (48, 61)
friday = (62, 75)

NAME = 'data_2.xlsx'

def find_cells(file_path, search_text, sheet_name='1 семестр'):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    found = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and (search_text == str(cell.value).strip()):  # Полное совпадение
                address = (get_column_letter(cell.column), cell.row)
                found.append(address)
    return found

def convert_address_to_indices(address):
    column_letter = ''
    row_number = ''
    for char in address:
        if char.isalpha():
                column_letter += char
        elif char.isdigit():
            row_number += char
    column_index = 0
    for i, char in enumerate(reversed(column_letter)):
        column_index += (ord(char.upper()) - ord('A') + 1) * (26 ** i)
    return (column_index, int(row_number))

# Загрузить книгу
wb = load_workbook('new_data.xlsx')
sheet = wb['Sheet1']

def read_range(sheet, start, end):
    values = []
    start_indices = convert_address_to_indices(start)
    end_indices = convert_address_to_indices(end)
    for row in sheet.iter_rows(min_col=start_indices[0], min_row=start_indices[1], max_col=end_indices[0], max_row=end_indices[1]):
        row_values = []
        for cell in row:
            if cell.value is not None:
                row_values.append(cell.value)
        if row_values:
            values.append(row_values)
    return values

def parse_data(group, week_day):
    if week_day == "monday":
        start = monday[0]
        end = monday[1]
    elif week_day == "tuesday":
        start = tuesday[0]
        end = tuesday[1]
    elif week_day == "wednesday":
        start = wednesday[0]
        end = wednesday[1]
    elif week_day == "thursday":
        start = thursday[0]
        end = thursday[1]
    elif week_day == "friday":
        start = friday[0]
        end = friday[1]
    group_letter = find_cells(NAME, f'{group}')
    if group_letter == []:
        return False, f"Группа {group} не найдена"
    data = read_range(sheet, f"{group_letter[0][0]}{start}", f"{group_letter[0][0]}{end}")
    return True, data